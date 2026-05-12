#!/usr/bin/env python3
"""简化测试版 - 只解析PET500"""
import os, json, re, csv
from datetime import datetime
import openpyxl
from qcloud_cos import CosConfig, CosS3Client

print("="*50)
print("🦐 测试PET500解析")
print(datetime.now().strftime("⏰ %Y-%m-%d %H:%M:%S"))
print("="*50)

# COS连接
config = CosConfig(Region='ap-beijing', 
    SecretId='AKIDjS3X7a2QgomdMExXgt0LYbS6ZqumU8aq',
    SecretKey='63wd7dGnCRVOQMKASbaQIzVqzzoSuCLQ')
client = CosS3Client(config)

def norm_date(s):
    from datetime import datetime as dt, timedelta
    if isinstance(s, dt):
        return s.strftime('%Y-%m-%d')
    s = str(s).strip()
    try:
        n = float(s)
        if 40000 < n < 50000:
            base = dt(1899, 12, 30)
            return (base + timedelta(days=n)).strftime('%Y-%m-%d')
    except ValueError:
        pass
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None

def pnum(s):
    try:
        return float(str(s).replace(',', ''))
    except:
        return 0.0

# 只处理PET500
sku_name = 'PET500_873480929689'
js_key = 'PET500'
prefix = f'data/{sku_name}/'

resp = client.list_objects(Bucket='tmall-bi-data-v1-1430009310', Prefix=prefix, Delimiter='/', MaxKeys=100)
subdirs = [p['Prefix'] for p in resp.get('CommonPrefixes', [])]
print(f'\n📦 {sku_name}: {len(subdirs)}个子文件夹')

all_dates = set()
sales_data = {}
ads_data = {}
refund_data = {}

for sub_prefix in subdirs:
    sub_name = sub_prefix.rstrip('/').split('/')[-1]
    resp2 = client.list_objects(Bucket='tmall-bi-data-v1-1430009310', Prefix=sub_prefix, MaxKeys=100)
    files = [f['Key'] for f in resp2.get('Contents', []) if not f['Key'].endswith('/')]
    
    print(f'  {sub_name}: {len(files)}个文件')
    
    for key in files:
        fname = os.path.basename(key)
        print(f'    - {fname}')
        
        # 下载
        r = client.get_object(Bucket='tmall-bi-data-v1-1430009310', Key=key)
        data = r['Body'].read()
        lpath = f'cos-downloads/{fname}'
        data = r['Body'].read()
        with open(lpath, 'wb') as f:
            f.write(data)
        
        # 解析Excel
        if fname.endswith(('.xlsx', '.xls')):
            try:
                wb = openpyxl.load_workbook(lpath, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    
                    # 找表头
                    header_idx = None
                    for i, row in enumerate(rows):
                        if row and any('日期' in str(c) for c in row if c):
                            header_idx = i
                            break
                    if header_idx is None:
                        print(f'      ⚠️ 没找到表头')
                        continue
                    
                    headers = [str(c).strip() if c else '' for c in rows[header_idx]]
                    print(f'      表头: {headers[:5]}...')
                    
                    for row in rows[header_idx+1:]:
                        if not row or not any(row):
                            continue
                        rd = dict(zip(headers, row))
                        
                        dc = next((h for h in headers if '日期' in h), None)
                        if not dc:
                            continue
                        dt = norm_date(rd.get(dc, ''))
                        if not dt:
                            continue
                        
                        # 存数据
                        if '销售' in sub_name or '销售' in fname:
                            if dt not in sales_data:
                                sales_data[dt] = {}
                            sales_data[dt].update(rd)
                            print(f'      ✅ 销售 {dt}: 支付金额={rd.get("支付金额", "N/A")}')
                        elif '投放' in sub_name or '推广' in fname:
                            if dt not in ads_data:
                                ads_data[dt] = {}
                            ads_data[dt].update(rd)
                            print(f'      ✅ 广告 {dt}: 花费={rd.get("花费", "N/A")}')
                        elif '退款' in sub_name or '售后' in fname:
                            if dt not in refund_data:
                                refund_data[dt] = {}
                            refund_data[dt].update(rd)
                            print(f'      ✅ 退款 {dt}: 退款率={rd.get("退款率", "N/A")}')
                        
                        all_dates.add(dt)
            
            except Exception as e:
                print(f'      ❌ 解析错误: {e}')

print(f"\n📅 日期范围: {min(all_dates)} ~ {max(all_dates)}")
print(f"   销售数据: {len(sales_data)}天")
print(f"   广告数据: {len(ads_data)}天")
print(f"   退款数据: {len(refund_data)}天")

# 生成data.json
dates_sorted = sorted(all_dates)
recent_14 = dates_sorted[-14:]

gmv = [pnum(sales_data.get(dt, {}).get('支付金额', 0)) / 10000 for dt in recent_14]
adSpend = [pnum(ads_data.get(dt, {}).get('花费', 0)) / 10000 for dt in recent_14]

print(f"\n📊 数据汇总:")
print(f"   GMV: {[f'{x:.1f}万' for x in gmv]}")
print(f"   广告花费: {[f'{x:.1f}万' for x in adSpend]}")

output = {
    "dates": recent_14,
    "skus": {
        js_key: {
            'name': sku_name,
            'gmv': [round(x, 2) for x in gmv],
            'adSpend': [round(x, 2) for x in adSpend],
            'net': [round(gmv[i] - pnum(refund_data.get(recent_14[i], {}).get('退款额', 0))/10000, 2) for i in range(len(recent_14))],
            'adRev': [round(pnum(ads_data.get(recent_14[i], {}).get('总成交金额', 0))/10000, 2) for i in range(len(recent_14))],
            'dir': [round(pnum(ads_data.get(recent_14[i], {}).get('直接成交金额', 0))/10000, 2) for i in range(len(recent_14))],
            'indir': [round(pnum(ads_data.get(recent_14[i], {}).get('间接成交金额', 0))/10000, 2) for i in range(len(recent_14))],
            'roi': [round(pnum(ads_data.get(recent_14[i], {}).get('总成交金额', 0))/pnum(ads_data.get(recent_14[i], {}).get('花费', 1)), 2) for i in range(len(recent_14))],
            'refund': [round(float(str(refund_data.get(recent_14[i], {}).get('退款率', '0')).replace('%',''))/100, 2) if '%' not in str(refund_data.get(recent_14[i], {}).get('退款率', '0')) else round(float(str(refund_data.get(recent_14[i], {}).get('退款率', '0')).replace('%','')), 2) for i in range(len(recent_14))],
            'paidTraf': [0.0]*len(recent_14),
            'advTraf': [0.0]*len(recent_14),
            'revisit': [0.0]*len(recent_14),
            'inner': [0.0]*len(recent_14),
        }
    },
    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M")
}

with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n🎉 生成 data.json 成功!")
print(f"   仪表盘: https://wg887k4tf4-lang.github.io/tmall-bi-dashboard/")
